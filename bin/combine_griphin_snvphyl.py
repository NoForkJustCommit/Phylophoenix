#!/usr/bin/env python3

#disable cache usage in the Python so __pycache__ isn't formed. If you don't do this using 'nextflow run cdcgov/phoenix...' a second time will causes and error
import sys
sys.dont_write_bytecode = True
import pandas as pd
import argparse
import glob
import os
import re
import xlsxwriter as ws
from xlsxwriter.utility import xl_rowcol_to_cell
import openpyxl
from openpyxl.styles import PatternFill, Font

# Function to get the script version
def get_version():
    return "1.0.0"

def parseArgs(args=None):
    parser = argparse.ArgumentParser(description="Add latitude and longitude to a dataset.")
    parser.add_argument("-g", "--griphin", required=True, help="Input GRiPHin_Summary.xlsx file")
    parser.add_argument('--version', action='version', version=get_version())# Add an argument to display the version
    return parser.parse_args()

def append_tsv_to_excel(old_griphin, snvmatrices, result_dict):
    # Load the original Excel file with openpyxl to preserve formatting and merged cells
    workbook = openpyxl.load_workbook(old_griphin)
    sheet = workbook.active
    # Determine the starting row for appending new data
    start_row = sheet.max_row + 2  # +2 to leave a blank line after existing data
    # Define a bold font for seq_type labels
    bold_font = openpyxl.styles.Font(bold=True)
    # Iterate through each TSV file to append
    for snvmatrix in snvmatrices:
        # Load the TSV file data
        snvmatrix_df = pd.read_csv(snvmatrix, sep='\t')
        # Write text, apply thistle color, and bold font in one line
        sheet.merge_cells("A" + str(start_row) + ":" + xl_rowcol_to_cell(start_row -1 , snvmatrix_df.shape[0]))
        snvmatrix_cell = sheet.cell(row=start_row, column=1, value="SNVPhyl Analysis: SNV Matrices")
        snvmatrix_cell.fill = PatternFill(start_color="D8BFD8", end_color="D8BFD8", fill_type="solid")
        snvmatrix_cell.font = Font(bold=True)
        # Derive seq_type from the filename by removing '_snvMatrix.tsv'
        seq_type = os.path.basename(snvmatrix).replace('_snvMatrix.tsv', '')
        # Write the seq_type label in bold
        sheet.cell(row=start_row + 1, column=1, value=seq_type).font = bold_font
        # Write the % core genome
        core = "% Core Genome Used: " + str(result_dict.get(seq_type))
        sheet.cell(row=start_row + 2, column=1, value=core)
        # Write the header of the TSV file below the seq_type label
        for col_idx, column_name in enumerate(snvmatrix_df.columns, start=1):
            #sheet.cell(row=start_row + 2, column=col_idx, value=column_name).font = bold_font
            sheet.cell(row=start_row + 4, column=col_idx, value=column_name)
        # Write the TSV data below the header
        for i, row in snvmatrix_df.iterrows():
            for j, value in enumerate(row):
                sheet.cell(row=start_row + i + 5, column=j + 1, value=value)  # +2 for seq_type and header rows
        # Update start_row to skip a line after the current TSV data
        start_row += len(snvmatrix_df) + 6  # +3 to leave one blank row after TSV data
    # Save the final output file
    workbook.save("SNVPhyl_GRiPHin_Summary.xlsx")
    print("Excel file with appended TSV data saved as 'SNVPhyl_GRiPHin_Summary.xlsx'.")

def get_sorted_files(pattern):
    # Retrieve all matching files
    files = glob.glob(pattern)
    # Separate files containing the All_STs and those that don't
    All_STs_files = [f for f in files if "All_STs" in f]
    other_files = [f for f in files if "All_STs" not in f]
    # Define function to extract numbers for sorting
    def extract_number(filename):
        match = re.search(r'\d+', filename)
        return int(match.group()) if match else float('inf')  # float('inf') puts files without numbers at the end
    # Sort remaining files by extracted number, then alphabetically for files without numbers
    other_files.sort(key=lambda f: (extract_number(f), f))
    # Concatenate lists with All_STs_files first
    return All_STs_files + other_files

def get_files():
    # You only need this for glob because glob will throw an index error if not.
    snvmatrices = get_sorted_files("*_snvMatrix.tsv")
    vcf2cores = get_sorted_files("*_vcf2core.tsv")
    #create empty dictionary to fill
    result_dict = {}
    # looping through vcf2core files
    for vcf2core in vcf2cores:
        # Derive seq_type from the filename by removing '_vcf2core.tsv'
        seq_type = os.path.basename(vcf2core).replace('_vcf2core.tsv', '')
        # Read the TSV file
        df_vcf2core = pd.read_csv(vcf2core, sep='\t')
        # Extract the last row's specified column value as a float - to get % core genome
        try:
            last_value = float(df_vcf2core["Percentage of all positions that are valid, included, and part of the core genome"].iloc[-1])
            result_dict[seq_type] = last_value
        except (KeyError, ValueError, IndexError) as e:
            print(f"Error processing file '{vcf2core}': {e}")
    return snvmatrices, result_dict

def main(old_griphin):
    snvmatrices, result_dict = get_files()
    append_tsv_to_excel(old_griphin, snvmatrices, result_dict)

if __name__ == "__main__":
    args = parseArgs()
    main(args.griphin)
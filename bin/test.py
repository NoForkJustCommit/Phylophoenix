#!/usr/bin/env python3

#disable cache usage in the Python so __pycache__ isn't formed. If you don't do this using 'nextflow run cdcgov/phoenix...' a second time will causes and error
import sys
sys.dont_write_bytecode = True
import pandas as pd
import argparse
import glob
import os
import xlsxwriter as ws
from xlsxwriter.utility import xl_rowcol_to_cell
import openpyxl

# Function to get the script version
def get_version():
    return "1.0.0"

def parseArgs(args=None):
    parser = argparse.ArgumentParser(description="Add latitude and longitude to a dataset.")
    parser.add_argument("-g", "--griphin", required=True, help="Input GRiPHin_Summary.xlsx file")
    parser.add_argument('--version', action='version', version=get_version())# Add an argument to display the version
    return parser.parse_args()

def append_tsv_to_excel(old_griphin, snvmatrices):
    # Load the original Excel file with openpyxl to preserve formatting and merged cells
    workbook = openpyxl.load_workbook(old_griphin)
    sheet = workbook.active
    # Determine the starting row for appending new data
    start_row = sheet.max_row + 2  # +2 to leave a blank line after existing data
    # Define a bold font for seq_type labels
    bold_font = openpyxl.styles.Font(bold=True)

    # Iterate through each TSV file to append
    for snvmatrix in snvmatrices:
        # Derive seq_type from the filename by removing '_snvMatrix.tsv'
        seq_type = os.path.basename(snvmatrix).replace('_snvMatrix.tsv', '')
        # Write the seq_type label in bold
        sheet.cell(row=start_row, column=1, value=seq_type).font = bold_font
        # Load the TSV file data
        snvmatrix_df = pd.read_csv(snvmatrix, sep='\t')
        # Write the TSV data below the seq_type label
        for i, row in snvmatrix_df.iterrows():
            for j, value in enumerate(row):
                sheet.cell(row=start_row + i + 1, column=j + 1, value=value)
        # Update start_row to skip a line after the current TSV data
        start_row += len(snvmatrix_df) + 3  # +3 to skip over TSV data and leave one blank row
    
    # Save the final output file
    workbook.save("GRiPHin_Summary.xlsx")

    # Set up Excel writer with xlsxwriter engine for formatting
    #with pd.ExcelWriter("GRiPHin_Summary.xlsx", engine='xlsxwriter') as writer:
       # Load the existing workbook into writer to keep the formatting
    #    writer.book = workbook
    #    writer.sheets = {ws.title: ws for ws in workbook.worksheets}
    #    worksheet = writer.sheets[sheet.title]
        ## Define bold format for the seq_type
        ##bold_format = writer.book.add_format({'bold': True})

        # Start row for appending new data after existing content
    #    start_row = len(griphin_df) + 2  # Leave a gap after existing data
        # Iterate through each TSV file to append
    #    for snvmatrix in snvmatrices:
            # Derive seq_type from the filename by removing '_snvTable.tsv'
    #        seq_type = os.path.basename(snvmatrix).replace('_snvMatrix', '')
            # Write the seq_type label in bold
    #        worksheet.write(start_row, 0, seq_type, bold_format)
            # Load the TSV file data
    #        snvmatrix_df = pd.read_csv(snvmatrix, sep='\t')
    #        print(snvmatrix_df)
            # Write the TSV data below the seq_type label
    #        snvmatrix_df.to_excel(writer, index=False, header=True, startrow=start_row + 1, sheet_name='Sheet1')
            # Update start_row to skip a line after the current TSV data
    #        start_row += len(snvmatrix_df) + 3  # +3 to skip over TSV data and leave one blank row
    print("Excel file with appended TSV data saved as 'GRiPHin_Summary.xlsx'.")


#def append_tsv_to_excel(excel_file, snvtables):
#    """ Appends a TSV file to the bottom of an Excel file without changing the existing Excel structure. """
#    # Load the Excel file into a DataFrame
#    excel_df = pd.read_excel(excel_file)
#    count = 0
#    for snvtable in snvtables:
#        # Derive seq_type from the filename by removing '_snvTable.tsv'
#        seq_type = os.path.basename(file_path).replace('_snvTable.tsv', '')
#        if count == 0:
#            # Load the TSV file into a DataFrame
#            snvtable_df = pd.read_csv(snvtable, sep='\t')
#            # Append the TSV DataFrame to the Excel DataFrame
#            combined_df = pd.concat([excel_df, snvtable_df], ignore_index=True)
#            count = count + 1
#        else:
#            combined_df = pd.concat([combined_df, snvtable_df], ignore_index=True)
    # Write the combined DataFrame back to a new Excel file
#    combined_df.to_excel("GRiPHin_Summary.xlsx", index=False)

def get_files():
    # You only need this for glob because glob will throw an index error if not.
    snvmatrices = glob.glob("*_snvMatrix.tsv")
    vcf2cores = glob.glob("*_vcf2core.tsv")
    #create empty dictionary to fill
    result_dict = {}
    # looping through vcf2core files
    for vcf2core in vcf2cores:
        # Derive seq_type from the filename by removing '_vcf2core.tsv'
        seq_type = os.path.basename(vcf2core).replace('_vcf2core.tsv', '')
        # Read the TSV file
        df_vcf2core = pd.read_csv(vcf2core, sep='\t')
        # Extract the last row's specified column value as a float - to get % core genome
        try:
            last_value = float(df_vcf2core["Percentage of all positions that are valid, included, and part of the core genome"].iloc[-1])
            result_dict[seq_type] = last_value
        except (KeyError, ValueError, IndexError) as e:
            print(f"Error processing file '{vcf2core}': {e}")
    print(result_dict)
    return snvmatrices, result_dict

def main(old_griphin):
    snvmatrices, result_dict = get_files()
    append_tsv_to_excel(old_griphin, snvmatrices)

if __name__ == "__main__":
    args = parseArgs()
    main(args.griphin)